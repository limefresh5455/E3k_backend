import re
from collections import OrderedDict
from pathlib import Path
from typing import Any, BinaryIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.supplier_addresses.schemas import SupplierCreate


SUPPLIER_NUMBER_RE = re.compile(r"^\d{6}$")
UPLOAD_COLUMNS = (
    "supplier_number",
    "name",
    "code_1",
    "code_2",
    "code_3",
    "code_4",
    "department_attention",
    "street",
    "postal_code",
    "city",
    "country",
    "business_phone",
    "private_phone",
    "mobile",
    "fax",
    "email",
)
SUPPLIER_UPLOAD_FIELDS = ("name", "code_1", "code_2", "code_3", "code_4")
ADDRESS_UPLOAD_FIELDS = UPLOAD_COLUMNS[6:]
MAX_UPLOAD_ROWS = 100_000
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def split_postal_city(value: Any) -> tuple[str | None, str | None]:
    text = _clean(value)
    if not text:
        return None, None
    parts = text.split(maxsplit=1)
    if len(parts) == 1:
        return parts[0], None
    return parts[0], parts[1].strip() or None


def parse_supplier_upload(source: BinaryIO) -> tuple[list[dict[str, Any]], int]:
    """Detect and parse either the client ERP report or the flat upload layout."""
    _validate_xlsx_archive(source)
    layout = _detect_upload_layout(source)
    if layout == "erp_report":
        return _parse_erp_supplier_upload(source)
    if layout == "flat":
        return _parse_flat_supplier_upload(source)
    raise ValueError(
        "Unsupported Excel layout. Expected the client address-list report containing A-Nr."
    )


def parse_supplier_master_upload(source: BinaryIO) -> tuple[list[dict[str, Any]], int]:
    """Parse only the four-row address-list format used by the canonical workbook."""
    _validate_xlsx_archive(source)
    if _detect_upload_layout(source) != "erp_report":
        raise ValueError(
            "Unsupported Excel layout. Upload an address-list workbook with the "
            "A-Nr. and Name/Firmenname headers."
        )
    return _parse_erp_supplier_upload(source)


def _detect_upload_layout(source: BinaryIO) -> str | None:
    source.seek(0)
    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        for row_number, row in enumerate(workbook.active.iter_rows(values_only=True), start=1):
            raw_headers = {(_clean(value) or "").lower() for value in row}
            normalized_headers = {_normalize_header(value) for value in row}
            if "supplier_number" in normalized_headers:
                return "flat"
            if "a-nr." in raw_headers and "name/firmenname" in raw_headers:
                return "erp_report"
            if row_number >= 30:
                break
    finally:
        workbook.close()
        source.seek(0)
    return None


def _parse_flat_supplier_upload(source: BinaryIO) -> tuple[list[dict[str, Any]], int]:
    """Parse a flat upload workbook where each row represents one address."""
    source.seek(0)
    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = None
        header_row_number = 0
        for header_row_number, candidate in enumerate(rows, start=1):
            if any(_clean(value) for value in candidate):
                header_row = candidate
                break
        if header_row is None:
            raise ValueError("Please upload data into the file and re-upload")

        headers = [_normalize_header(value) for value in header_row]
        duplicates = sorted({header for header in headers if header and headers.count(header) > 1})
        if duplicates:
            raise ValueError(f"Duplicate Excel columns: {', '.join(duplicates)}")
        missing = [column for column in UPLOAD_COLUMNS if column not in headers]
        if missing:
            raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")
        indexes = {column: headers.index(column) for column in UPLOAD_COLUMNS}

        grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
        rows_read = 0
        for excel_row_number, row in enumerate(rows, start=header_row_number + 1):
            if not any(_clean(value) for value in row):
                continue
            rows_read += 1
            if rows_read > MAX_UPLOAD_ROWS:
                raise ValueError(f"Excel file exceeds the {MAX_UPLOAD_ROWS:,}-row limit")
            supplier_number = _normalize_supplier_number(_value_at(row, indexes["supplier_number"]))
            if not supplier_number:
                raise ValueError(f"Row {excel_row_number}: supplier_number is required")

            incoming_supplier = {
                field: _clean(_value_at(row, indexes[field])) for field in SUPPLIER_UPLOAD_FIELDS
            }
            address = {
                field: _clean(_value_at(row, indexes[field])) for field in ADDRESS_UPLOAD_FIELDS
            }
            group = grouped.setdefault(
                supplier_number,
                {
                    "supplier_number": supplier_number,
                    **incoming_supplier,
                    "addresses": [],
                    "_address_keys": set(),
                },
            )
            _merge_supplier_fields(group, incoming_supplier, supplier_number, excel_row_number)
            address_key = tuple(address[field] for field in ADDRESS_UPLOAD_FIELDS)
            if address_key not in group["_address_keys"]:
                group["addresses"].append(address)
                group["_address_keys"].add(address_key)
    finally:
        workbook.close()

    records = []
    if not grouped:
        raise ValueError("Please upload data into the file and re-upload")
    for supplier_number, group in grouped.items():
        group.pop("_address_keys")
        records.append(SupplierCreate.model_validate(group).model_dump())
    return records, rows_read


def _parse_erp_supplier_upload(source: BinaryIO) -> tuple[list[dict[str, Any]], int]:
    """Parse the client's four-line ERP address-list report."""
    source.seek(0)
    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
    records_read = 0
    physical_rows = 0
    pending_block: list[tuple[Any, ...]] = []
    pending_start_row = 0
    try:
        for row in workbook.active.iter_rows(values_only=True):
            physical_rows += 1
            _check_row_limit(physical_rows)
            supplier_number = _normalize_supplier_number(_value_at(row, 0))
            is_supplier_start = bool(
                supplier_number and SUPPLIER_NUMBER_RE.fullmatch(supplier_number)
            )

            if is_supplier_start:
                if pending_block:
                    _add_erp_block(grouped, pending_block, pending_start_row)
                    records_read += 1
                pending_block = [row]
                pending_start_row = physical_rows
                continue
            if pending_block and len(pending_block) < 4:
                pending_block.append(row)
                if len(pending_block) == 4:
                    _add_erp_block(grouped, pending_block, pending_start_row)
                    records_read += 1
                    pending_block = []
        if pending_block:
            _add_erp_block(grouped, pending_block, pending_start_row)
            records_read += 1
    finally:
        workbook.close()
        source.seek(0)

    if not grouped:
        raise ValueError("Please upload data into the file and re-upload")
    records = []
    for group in grouped.values():
        group.pop("_address_keys")
        records.append(SupplierCreate.model_validate(group).model_dump())
    return records, records_read


def _add_erp_block(
    grouped: OrderedDict[str, dict[str, Any]], block: list[tuple[Any, ...]], start_row: int
) -> None:
    block = [*block, *([()] * (4 - len(block)))]
    supplier_number = _normalize_supplier_number(_value_at(block[0], 0))
    if supplier_number is None:
        return
    postal_code, city = split_postal_city(_value_at(block[3], 1))
    incoming_supplier = {
        "name": _clean(_value_at(block[0], 1)),
        "code_1": _clean(_value_at(block[0], 7)),
        "code_2": _clean(_value_at(block[0], 8)),
        "code_3": _clean(_value_at(block[1], 7)),
        "code_4": _clean(_value_at(block[1], 8)),
    }
    address = {
        "department_attention": _clean(_value_at(block[1], 1)),
        "street": _clean(_value_at(block[2], 1)),
        "postal_code": postal_code,
        "city": city, 
        "country": _clean(_value_at(block[3], 4)),
        "business_phone": _clean(_value_at(block[0], 5)),
        "private_phone": _clean(_value_at(block[0], 6)),
        "mobile": _clean(_value_at(block[1], 5)),
        "fax": _clean(_value_at(block[1], 6)),
        "email": _clean(_value_at(block[2], 5)),
    }
    group = grouped.setdefault(
        supplier_number,
        {
            "supplier_number": supplier_number,
            **incoming_supplier,
            "addresses": [],
            "_address_keys": set(),
        },
    )
    _merge_supplier_fields(group, incoming_supplier, supplier_number, start_row)
    address_key = tuple(address[field] for field in ADDRESS_UPLOAD_FIELDS)
    if address_key not in group["_address_keys"]:
        group["addresses"].append(address)
        group["_address_keys"].add(address_key)


def _check_row_limit(row_count: int) -> None:
    if row_count > MAX_UPLOAD_ROWS:
        raise ValueError(f"Excel file exceeds the {MAX_UPLOAD_ROWS:,}-row limit")


def _validate_xlsx_archive(source: BinaryIO) -> None:
    """Reject archive expansion attacks before openpyxl processes the workbook."""
    source.seek(0)
    try:
        with ZipFile(source) as archive:
            uncompressed_size = sum(item.file_size for item in archive.infolist())
            if uncompressed_size > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("Expanded Excel content exceeds the 250 MB safety limit")
    except BadZipFile:
        raise
    finally:
        source.seek(0)


def _normalize_header(value: Any) -> str:
    return (_clean(value) or "").lower().replace(" ", "_").replace("-", "_")


def _normalize_supplier_number(value: Any) -> str | None:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = _clean(value)
    if text and text.isdigit() and len(text) <= 6:
        return text.zfill(6)
    return text


def _value_at(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _merge_supplier_fields(
    group: dict[str, Any], incoming: dict[str, str | None], supplier_number: str, row_number: int
) -> None:
    for field in SUPPLIER_UPLOAD_FIELDS:
        value = incoming[field]
        current = group.get(field)
        if current is None and value is not None:
            group[field] = value
        elif current is not None and value is not None and current != value:
            raise ValueError(
                f"Row {row_number}: conflicting {field} for supplier {supplier_number}"
            )


def parse_supplier_workbook(path: str | Path) -> list[dict[str, Any]]:
    """Parse a supported supplier workbook from a filesystem path."""
    with Path(path).open("rb") as source:
        records, _ = parse_supplier_upload(source)
    return records
