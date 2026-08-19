import logging
import os
import shutil
import time
from contextlib import contextmanager
from copy import copy
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Iterator
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import SUPPLIER_WORKBOOK_PATH, SUPPLIER_WORKBOOK_TEMPLATE_PATH
from app.supplier_addresses.errors import SupplierConflictError, SupplierWorkbookError


logger = logging.getLogger("supplier_excel_export")
BLOCK_HEIGHT = 4
STYLE_SOURCE_START_ROW = 15
LOCK_TIMEOUT_SECONDS = 15.0
STALE_LOCK_SECONDS = 120.0


def get_master_workbook_path() -> Path:
    """Return the initialized, downloadable supplier master workbook."""
    _ensure_master_workbook()
    return SUPPLIER_WORKBOOK_PATH


def append_supplier_to_master_workbook(supplier: dict[str, Any]) -> Path:
    """Append one newly created supplier in the report's four-row layout."""
    result = append_suppliers_to_master_workbook([supplier], skip_existing=False)
    return result["path"]


def append_suppliers_to_master_workbook(
    suppliers: list[dict[str, Any]], *, skip_existing: bool = True
) -> dict[str, Any]:
    """Atomically append multiple suppliers to the canonical workbook."""
    if not suppliers:
        return {"path": get_master_workbook_path(), "appended": 0, "already_present": 0}

    _ensure_master_workbook()
    with _exclusive_file_lock(SUPPLIER_WORKBOOK_PATH):
        workbook = None
        temporary_path: Path | None = None
        try:
            workbook = load_workbook(SUPPLIER_WORKBOOK_PATH)
            worksheet = workbook.active
            existing_numbers = _workbook_supplier_numbers(worksheet)
            appended = 0
            already_present = 0
            address_blocks = 0
            for supplier in suppliers:
                supplier_number = _supplier_number(supplier.get("supplier_number"))
                if not supplier_number:
                    raise SupplierWorkbookError(
                        "Cannot append a supplier without a supplier number"
                    )
                if supplier_number in existing_numbers:
                    if not skip_existing:
                        raise SupplierConflictError(
                            f"Supplier {supplier_number} already exists in the Excel workbook"
                        )
                    already_present += 1
                    continue

                addresses = supplier.get("addresses") or [{}]
                for address in addresses:
                    _append_address_block(worksheet, supplier, address, supplier_number)
                    address_blocks += 1
                existing_numbers.add(supplier_number)
                appended += 1

            if appended == 0:
                workbook.close()
                workbook = None
                return {
                    "path": SUPPLIER_WORKBOOK_PATH,
                    "appended": 0,
                    "already_present": already_present,
                }

            SUPPLIER_WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
            with NamedTemporaryFile(
                prefix="supplier-addresses-",
                suffix=".xlsx",
                dir=SUPPLIER_WORKBOOK_PATH.parent,
                delete=False,
            ) as temporary:
                temporary_path = Path(temporary.name)

            workbook.save(temporary_path)
            workbook.close()
            workbook = None

            # Validate the complete archive before replacing the downloadable master.
            validation_workbook = load_workbook(temporary_path, read_only=True, data_only=False)
            validation_workbook.close()
            os.replace(temporary_path, SUPPLIER_WORKBOOK_PATH)
            temporary_path = None
            logger.info(
                "Suppliers appended to master workbook: appended=%d, already_present=%d, "
                "address_blocks=%d",
                appended,
                already_present,
                address_blocks,
            )
            return {
                "path": SUPPLIER_WORKBOOK_PATH,
                "appended": appended,
                "already_present": already_present,
            }
        except SupplierConflictError:
            raise
        except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
            raise SupplierWorkbookError(
                f"Could not update supplier Excel workbook: {exc}"
            ) from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)


def _ensure_master_workbook() -> None:
    if SUPPLIER_WORKBOOK_PATH.is_file():
        return
    if not SUPPLIER_WORKBOOK_TEMPLATE_PATH.is_file():
        raise SupplierWorkbookError(
            f"Supplier workbook template was not found: {SUPPLIER_WORKBOOK_TEMPLATE_PATH}"
        )

    SUPPLIER_WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _exclusive_file_lock(SUPPLIER_WORKBOOK_PATH):
        if SUPPLIER_WORKBOOK_PATH.is_file():
            return
        try:
            shutil.copy2(SUPPLIER_WORKBOOK_TEMPLATE_PATH, SUPPLIER_WORKBOOK_PATH)
        except OSError as exc:
            raise SupplierWorkbookError(
                f"Could not initialize supplier Excel workbook: {exc}"
            ) from exc


def _append_address_block(worksheet, supplier, address, supplier_number: str) -> None:
    start_row = worksheet.max_row + 1
    _copy_block_format(worksheet, start_row)

    worksheet.cell(start_row, 1, supplier_number)
    worksheet.cell(start_row, 1).number_format = "@"
    worksheet.cell(start_row, 2, supplier.get("name"))
    worksheet.cell(start_row, 6, address.get("business_phone"))
    worksheet.cell(start_row, 7, address.get("private_phone"))
    worksheet.cell(start_row, 8, supplier.get("code_1"))
    worksheet.cell(start_row, 9, supplier.get("code_2"))

    worksheet.cell(start_row + 1, 2, address.get("department_attention"))
    worksheet.cell(start_row + 1, 6, address.get("mobile"))
    worksheet.cell(start_row + 1, 7, address.get("fax"))
    worksheet.cell(start_row + 1, 8, supplier.get("code_3"))
    worksheet.cell(start_row + 1, 9, supplier.get("code_4"))

    worksheet.cell(start_row + 2, 2, address.get("street"))
    worksheet.cell(start_row + 2, 6, address.get("email"))

    worksheet.cell(
        start_row + 3,
        2,
        _postal_city(address.get("postal_code"), address.get("city")),
    )
    worksheet.cell(start_row + 3, 5, address.get("country"))


def _copy_block_format(worksheet, target_start_row: int) -> None:
    if worksheet.max_row < STYLE_SOURCE_START_ROW + BLOCK_HEIGHT - 1:
        return
    for offset in range(BLOCK_HEIGHT):
        source_row = STYLE_SOURCE_START_ROW + offset
        target_row = target_start_row + offset
        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in range(1, worksheet.max_column + 1):
            source = worksheet.cell(source_row, column)
            target = worksheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)


def _workbook_supplier_numbers(worksheet) -> set[str]:
    return {
        normalized
        for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True)
        if (normalized := _supplier_number(row[0]))
    }


def _supplier_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return f"{value:06d}"
    if isinstance(value, float) and value.is_integer():
        return f"{int(value):06d}"
    text = str(value).strip()
    return text.zfill(6) if text.isdigit() and len(text) < 6 else text


def _postal_city(postal_code: Any, city: Any) -> str | None:
    parts = [str(value).strip() for value in (postal_code, city) if value is not None and str(value).strip()]
    return "  ".join(parts) or None


@contextmanager
def _exclusive_file_lock(workbook_path: Path) -> Iterator[None]:
    lock_path = workbook_path.with_suffix(f"{workbook_path.suffix}.lock")
    deadline = time.monotonic() + LOCK_TIMEOUT_SECONDS
    lock_fd: int | None = None
    while lock_fd is None:
        try:
            lock_fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(lock_fd, f"{os.getpid()}\n".encode("ascii"))
        except FileExistsError:
            try:
                if time.time() - lock_path.stat().st_mtime > STALE_LOCK_SECONDS:
                    lock_path.unlink(missing_ok=True)
                    continue
            except FileNotFoundError:
                continue
            if time.monotonic() >= deadline:
                raise SupplierWorkbookError("Supplier Excel workbook is busy; please retry")
            time.sleep(0.1)
    try:
        yield
    finally:
        if lock_fd is not None:
            os.close(lock_fd)
        lock_path.unlink(missing_ok=True)
